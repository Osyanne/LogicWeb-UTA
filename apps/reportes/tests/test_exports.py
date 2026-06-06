from django.test import TestCase

from apps.ejercicios.models import Usuario, Tema, Ejercicio, Intento
from apps.reportes import services, exporters


def _data(user):
    return services.progreso_estudiante(user)


class ExcelExporterTest(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(
            username='est', password='clave12345', first_name='Ana', last_name='Pérez',
        )
        tema = Tema.objects.create(nombre_tema='Algoritmos', descripcion='x', unidad=1, orden=1)
        ej = Ejercicio.objects.create(
            titulo='Suma', enunciado='e', categoria='interactivo', tema=tema,
            codigo_cpp='x', solucion_esperada='5', tipo_respuesta='entero',
        )
        Intento.objects.create(usuario=self.user, ejercicio=ej, respuesta_usuario='5', resultado='correcto')

    def test_excel_es_xlsx(self):
        contenido = exporters.reporte_estudiante_excel(_data(self.user), self.user)
        self.assertIsInstance(contenido, bytes)
        self.assertTrue(contenido.startswith(b'PK'))     # magic de zip/xlsx
        self.assertGreater(len(contenido), 1000)

    def test_excel_estudiante_sin_intentos(self):
        vacio = Usuario.objects.create_user(username='vacio', password='clave12345')
        contenido = exporters.reporte_estudiante_excel(_data(vacio), vacio)
        self.assertTrue(contenido.startswith(b'PK'))


class PdfExporterTest(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(
            username='est', password='clave12345', first_name='Ana', last_name='Pérez',
        )
        tema = Tema.objects.create(nombre_tema='Algoritmos', descripcion='x', unidad=1, orden=1)
        ej = Ejercicio.objects.create(
            titulo='Suma', enunciado='e', categoria='interactivo', tema=tema,
            codigo_cpp='x', solucion_esperada='5', tipo_respuesta='entero',
        )
        Intento.objects.create(usuario=self.user, ejercicio=ej, respuesta_usuario='5', resultado='correcto')

    def test_pdf_es_pdf(self):
        contenido = exporters.reporte_estudiante_pdf(_data(self.user), self.user)
        self.assertIsInstance(contenido, bytes)
        self.assertTrue(contenido.startswith(b'%PDF'))
        self.assertGreater(len(contenido), 1000)

    def test_pdf_estudiante_sin_intentos(self):
        vacio = Usuario.objects.create_user(username='vacio', password='clave12345')
        contenido = exporters.reporte_estudiante_pdf(_data(vacio), vacio)
        self.assertTrue(contenido.startswith(b'%PDF'))

    def test_pdf_titulo_largo_no_rompe(self):
        tema = Tema.objects.create(nombre_tema='T' * 100, descripcion='x', unidad=2, orden=1)
        ej = Ejercicio.objects.create(
            titulo='X' * 200, enunciado='e', categoria='resuelto', tema=tema,
            codigo_cpp='x', solucion_esperada='1', tipo_respuesta='entero',
        )
        Intento.objects.create(usuario=self.user, ejercicio=ej, respuesta_usuario='1', resultado='correcto')
        contenido = exporters.reporte_estudiante_pdf(_data(self.user), self.user)
        self.assertTrue(contenido.startswith(b'%PDF'))
        self.assertGreater(len(contenido), 1000)


class ExcelContenidoTest(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(
            username='ana', password='clave12345', first_name='Ana', last_name='Pérez',
        )
        tema = Tema.objects.create(nombre_tema='Algoritmos', descripcion='x', unidad=1, orden=1)
        ej = Ejercicio.objects.create(
            titulo='Suma', enunciado='e', categoria='interactivo', tema=tema,
            codigo_cpp='x', solucion_esperada='5', tipo_respuesta='entero',
        )
        Intento.objects.create(usuario=self.user, ejercicio=ej, respuesta_usuario='5', resultado='correcto')

    def test_excel_hojas_y_contenido(self):
        import io
        from openpyxl import load_workbook
        contenido = exporters.reporte_estudiante_excel(_data(self.user), self.user)
        wb = load_workbook(io.BytesIO(contenido))
        self.assertEqual(wb.sheetnames, ['Resumen', 'Historial'])
        hist = wb['Historial']
        encabezado = [c.value for c in hist[1]]
        self.assertEqual(
            encabezado,
            ['Ejercicio', 'Unidad', 'Tema', 'Tipo', 'Tu respuesta', 'Resultado', 'Fecha'],
        )
        self.assertEqual(hist.cell(row=2, column=1).value, 'Suma')
        self.assertEqual(hist.cell(row=2, column=2).value, 'U1')
        self.assertEqual(hist.cell(row=2, column=6).value, 'correcto')


from django.urls import reverse


class ExportEndpointTest(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(username='est', password='clave12345')

    def test_pdf_requiere_login(self):
        self.assertEqual(self.client.get(reverse('exportar_pdf')).status_code, 302)

    def test_excel_requiere_login(self):
        self.assertEqual(self.client.get(reverse('exportar_excel')).status_code, 302)

    def test_pdf_descarga(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('exportar_pdf'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertIn('attachment', resp['Content-Disposition'])
        self.assertIn('.pdf', resp['Content-Disposition'])
        self.assertTrue(resp.content.startswith(b'%PDF'))

    def test_excel_descarga(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('exportar_excel'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', resp['Content-Disposition'])
        self.assertTrue(resp.content.startswith(b'PK'))

    def test_csv_descarga(self):
        tema = Tema.objects.create(nombre_tema='Algoritmos', descripcion='x', unidad=1, orden=1)
        ej = Ejercicio.objects.create(
            titulo='SumaCSV', enunciado='e', categoria='interactivo', tema=tema,
            codigo_cpp='x', solucion_esperada='5', tipo_respuesta='entero',
        )
        Intento.objects.create(usuario=self.user, ejercicio=ej, respuesta_usuario='5', resultado='correcto')
        self.client.force_login(self.user)
        resp = self.client.get(reverse('exportar_csv'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'text/csv')
        self.assertIn('attachment', resp['Content-Disposition'])
        self.assertContains(resp, 'SumaCSV')   # la fila del intento aparece en el CSV
