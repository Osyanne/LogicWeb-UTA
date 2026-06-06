"""Exportadores del reporte de progreso del estudiante a Excel y PDF.

Reciben el dict de services.progreso_estudiante(user) y el user, y devuelven bytes.
No tocan HTTP: la vista envuelve los bytes en HttpResponse.
"""
import io

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


AZUL_UTA = colors.HexColor('#4a90e2')


def _nombre(user):
    return user.get_full_name() or user.username


def _estilo_encabezado(ws, fila, ncols):
    relleno = PatternFill('solid', fgColor='4A90E2')
    for col in range(1, ncols + 1):
        c = ws.cell(row=fila, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = relleno
        c.alignment = Alignment(horizontal='center')


def _autoancho(ws):
    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ancho + 4, 50)


def reporte_estudiante_excel(data: dict, user) -> bytes:
    wb = Workbook()

    # — Hoja Resumen —
    ws = wb.active
    ws.title = 'Resumen'
    ws.append([f'Reporte de progreso — {_nombre(user)}'])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F4E79')

    ws.append([])
    ws.append(['Métrica', 'Valor'])
    _estilo_encabezado(ws, ws.max_row, 2)
    ws.append(['Aciertos', data['correctos']])
    ws.append(['Errores', data['incorrectos']])
    ws.append(['Total de intentos', data['total']])
    ws.append(['Porcentaje de éxito', f"{data['porcentaje']}%"])
    ws.append(['Ejercicios estudiados', data['codigos_vistos']])

    ws.append([])
    fila_cab = ws.max_row + 1
    ws.append(['Unidad', 'Total', 'Correctos', 'Porcentaje'])
    _estilo_encabezado(ws, fila_cab, 4)
    for u in data['progreso_unidades']:
        ws.append([u['nombre'], u['total'], u['correctos'], f"{u['porcentaje']}%"])
    _autoancho(ws)

    # — Hoja Historial —
    hist = wb.create_sheet('Historial')
    hist.append(['Ejercicio', 'Unidad', 'Tema', 'Tipo', 'Tu respuesta', 'Resultado', 'Fecha'])
    _estilo_encabezado(hist, 1, 7)
    for i in data['intentos']:
        hist.append([
            i['titulo'],
            f"U{i['unidad']}",
            i['tema'],
            i['categoria_display'],
            i['respuesta'],
            i['resultado'],
            i['fecha'].strftime('%d/%m/%Y %H:%M'),
        ])
    hist.freeze_panes = 'A2'
    hist.auto_filter.ref = f"A1:G{max(hist.max_row, 1)}"
    _autoancho(hist)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _estilo_tabla():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), AZUL_UTA),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef3fb')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ])


def _pie(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#888888'))
    canvas.drawString(2 * cm, 1.2 * cm, 'LogicWeb UTA — Reporte generado automáticamente')
    canvas.drawRightString(19 * cm, 1.2 * cm, f'Página {doc.page}')
    canvas.restoreState()


def reporte_estudiante_pdf(data: dict, user) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm, leftMargin=2 * cm, rightMargin=2 * cm,
        title='Reporte de progreso — LogicWeb UTA',
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('TituloUTA', parent=estilos['Title'],
                                   fontSize=16, textColor=AZUL_UTA, spaceAfter=4)
    estilo_sub = ParagraphStyle('Sub', parent=estilos['Normal'],
                                fontSize=10, textColor=colors.HexColor('#555555'))
    estilo_h2 = ParagraphStyle('H2UTA', parent=estilos['Heading2'],
                               fontSize=12, textColor=AZUL_UTA, spaceBefore=14, spaceAfter=6)

    el = []
    el.append(Paragraph('UNIVERSIDAD TÉCNICA DE AMBATO', estilo_titulo))
    el.append(Paragraph('LogicWeb UTA — Reporte de progreso', estilo_sub))
    el.append(Paragraph(f'Estudiante: {_nombre(user)}', estilo_sub))
    el.append(Paragraph(f'Generado: {timezone.localtime():%d/%m/%Y %H:%M}', estilo_sub))
    el.append(Spacer(1, 0.4 * cm))

    # KPIs
    el.append(Paragraph('Resumen', estilo_h2))
    kpis = [
        ['Aciertos', 'Errores', 'Total', '% Éxito', 'Estudiados'],
        [str(data['correctos']), str(data['incorrectos']), str(data['total']),
         f"{data['porcentaje']}%", str(data['codigos_vistos'])],
    ]
    t_kpis = Table(kpis, hAlign='LEFT')
    t_kpis.setStyle(_estilo_tabla())
    el.append(t_kpis)

    # Progreso por unidad
    if data['progreso_unidades']:
        el.append(Paragraph('Progreso por unidad', estilo_h2))
        filas = [['Unidad', 'Total', 'Correctos', '%']]
        for u in data['progreso_unidades']:
            filas.append([u['nombre'], str(u['total']), str(u['correctos']), f"{u['porcentaje']}%"])
        t = Table(filas, hAlign='LEFT', colWidths=[9 * cm, 2.5 * cm, 2.5 * cm, 2 * cm])
        t.setStyle(_estilo_tabla())
        el.append(t)

    # Historial
    el.append(Paragraph('Historial de ejercicios', estilo_h2))
    if data['intentos']:
        filas = [['Ejercicio', 'Unidad / Tema', 'Tipo', 'Resultado', 'Fecha']]
        for i in data['intentos']:
            filas.append([
                i['titulo'],
                f"U{i['unidad']} · {i['tema']}",
                i['categoria_display'],
                i['resultado'],
                i['fecha'].strftime('%d/%m/%Y %H:%M'),
            ])
        t = Table(filas, hAlign='LEFT', repeatRows=1,
                  colWidths=[4.5 * cm, 4.5 * cm, 2.8 * cm, 2.4 * cm, 3 * cm])
        t.setStyle(_estilo_tabla())
        el.append(t)
    else:
        el.append(Paragraph('Aún no has practicado ningún ejercicio.', estilo_sub))

    doc.build(el, onFirstPage=_pie, onLaterPages=_pie)
    return buffer.getvalue()
